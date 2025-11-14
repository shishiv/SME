#!/usr/bin/env python3
"""
Script para criar uma planilha Excel com dashboard de tutoriais
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_tutorial_dashboard():
    """Cria a planilha dashboard com múltiplas abas de tutorial"""

    wb = Workbook()
    wb.remove(wb.active)  # Remove a aba padrão

    # Cores para estilo
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. Criar aba Dashboard/Índice
    ws_dashboard = wb.create_sheet("📊 Dashboard", 0)

    ws_dashboard['A1'] = "DASHBOARD DE TUTORIAIS"
    ws_dashboard['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_dashboard['A1'].fill = title_fill
    ws_dashboard.merge_cells('A1:D1')
    ws_dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_dashboard.row_dimensions[1].height = 30

    # Data de criação
    ws_dashboard['A2'] = f"Data de Atualização: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_dashboard['A2'].font = Font(italic=True, size=10)
    ws_dashboard.merge_cells('A2:D2')

    # Lista de tutoriais disponíveis
    ws_dashboard['A4'] = "TUTORIAIS DISPONÍVEIS"
    ws_dashboard['A4'].font = section_font
    ws_dashboard['A4'].fill = section_fill
    ws_dashboard.merge_cells('A4:D4')

    tutorials = [
        ("Introdução ao Excel", "Conceitos básicos de planilhas"),
        ("Fórmulas Essenciais", "Operações matemáticas e lógicas"),
        ("Formatação Avançada", "Estilos, cores e layouts"),
        ("Gráficos e Visualização", "Criar gráficos e dashboards"),
        ("Análise de Dados", "Pivot tables e análise estatística"),
        ("Automação com VBA", "Macros e scripts básicos"),
    ]

    headers = ["#", "Tutorial", "Descrição", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws_dashboard.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row_idx, (title, desc) in enumerate(tutorials, 6):
        ws_dashboard.cell(row=row_idx, column=1).value = row_idx - 5
        ws_dashboard.cell(row=row_idx, column=2).value = title
        ws_dashboard.cell(row=row_idx, column=3).value = desc
        ws_dashboard.cell(row=row_idx, column=4).value = "✓ Completo"

        for col in range(1, 5):
            cell = ws_dashboard.cell(row=row_idx, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Ajustar largura das colunas
    ws_dashboard.column_dimensions['A'].width = 5
    ws_dashboard.column_dimensions['B'].width = 25
    ws_dashboard.column_dimensions['C'].width = 35
    ws_dashboard.column_dimensions['D'].width = 15

    # 2. Tutorial 1: Introdução ao Excel
    ws_intro = wb.create_sheet("1️⃣ Introdução", 1)
    _create_tutorial_tab(ws_intro, "Introdução ao Excel", [
        ("O que é uma Planilha?", "Uma planilha é um documento eletrônico composto por linhas e colunas que formam células, usadas para armazenar e manipular dados."),
        ("Componentes Básicos", [
            "• Células: Interseção de linhas e colunas (A1, B2, etc)",
            "• Linhas: Identificadas por números (1, 2, 3...)",
            "• Colunas: Identificadas por letras (A, B, C...)",
            "• Abas: Folhas diferentes dentro do mesmo arquivo",
        ]),
        ("Como Começar", [
            "1. Abrir o Excel ou similar",
            "2. Criar um novo arquivo",
            "3. Clicar em uma célula",
            "4. Digitar dados ou fórmulas",
        ]),
    ], header_fill, header_font, section_fill, section_font, border)

    # 3. Tutorial 2: Fórmulas Essenciais
    ws_formulas = wb.create_sheet("2️⃣ Fórmulas", 2)
    _create_tutorial_tab(ws_formulas, "Fórmulas Essenciais", [
        ("SUM - Soma", [
            "Sintaxe: =SUM(intervalo)",
            "Exemplo: =SUM(A1:A10) → soma de A1 até A10",
            "Uso: Calcular totais",
        ]),
        ("AVERAGE - Média", [
            "Sintaxe: =AVERAGE(intervalo)",
            "Exemplo: =AVERAGE(B1:B5) → média dos valores",
            "Uso: Encontrar o valor médio",
        ]),
        ("IF - Condicional", [
            "Sintaxe: =IF(condição, valor_verdadeiro, valor_falso)",
            "Exemplo: =IF(A1>10, 'Aprovado', 'Reprovado')",
            "Uso: Lógica condicional",
        ]),
        ("COUNTIF - Contar com Critério", [
            "Sintaxe: =COUNTIF(intervalo, critério)",
            "Exemplo: =COUNTIF(A1:A10, '>5')",
            "Uso: Contar células que atendem um critério",
        ]),
    ], header_fill, header_font, section_fill, section_font, border)

    # 4. Tutorial 3: Formatação Avançada
    ws_format = wb.create_sheet("3️⃣ Formatação", 3)
    _create_tutorial_tab(ws_format, "Formatação Avançada", [
        ("Formatação de Células", [
            "1. Selecionar as células",
            "2. Clique direito → Formatar Células",
            "3. Escolha: Número, Moeda, Percentual, etc",
            "4. Aplicar",
        ]),
        ("Cores e Fundos", [
            "• Cor de Fundo: Menu Início → Cor de Preenchimento",
            "• Cor da Fonte: Botão de cor da fonte",
            "• Bordas: Formatar Células → Aba Bordas",
        ]),
        ("Estilos Predefinidos", [
            "• Excel oferece estilos prontos",
            "• Menu Início → Estilos",
            "• Escolha um tema e aplique",
        ]),
    ], header_fill, header_font, section_fill, section_font, border)

    # 5. Tutorial 4: Gráficos e Visualização
    ws_charts = wb.create_sheet("4️⃣ Gráficos", 4)
    _create_tutorial_tab(ws_charts, "Gráficos e Visualização", [
        ("Tipos de Gráficos", [
            "• Colunas: Comparar valores",
            "• Linhas: Mostrar tendências",
            "• Pizza: Mostrar proporções",
            "• Barras: Comparar horizontalmente",
        ]),
        ("Criar um Gráfico", [
            "1. Selecionar dados (incluindo rótulos)",
            "2. Menu Inserir → Gráfico",
            "3. Escolher tipo de gráfico",
            "4. Configurar eixos e títulos",
            "5. Inserir",
        ]),
        ("Dicas de Design", [
            "• Use cores contrastantes",
            "• Adicione títulos descritivos",
            "• Inclua legendas claras",
            "• Evite gráficos muito complexos",
        ]),
    ], header_fill, header_font, section_fill, section_font, border)

    # 6. Tutorial 5: Análise de Dados
    ws_analysis = wb.create_sheet("5️⃣ Análise de Dados", 5)
    _create_tutorial_tab(ws_analysis, "Análise de Dados", [
        ("Pivot Table (Tabela Dinâmica)", [
            "1. Selecionar dados",
            "2. Menu Inserir → Tabela Dinâmica",
            "3. Escolher campo para linhas/colunas",
            "4. Adicionar valores a somar/contar",
            "5. Criar",
        ]),
        ("Filtros e Classificação", [
            "• Filtro Automático: Dados → Autofilter",
            "• Ordenar: Dados → Classificar",
            "• Filtro Avançado: Dados → Filtro Avançado",
        ]),
        ("Funções Estatísticas", [
            "• STDEV: Desvio padrão",
            "• VAR: Variância",
            "• MIN/MAX: Mínimo e máximo",
            "• PERCENTILE: Percentil",
        ]),
    ], header_fill, header_font, section_fill, section_font, border)

    # 7. Tutorial 6: Automação com VBA
    ws_vba = wb.create_sheet("6️⃣ Automação", 6)
    _create_tutorial_tab(ws_vba, "Automação com VBA", [
        ("O que é VBA?", "VBA (Visual Basic for Applications) é uma linguagem de programação que permite automatizar tarefas no Excel."),
        ("Acessar o Editor VBA", [
            "1. Pressionar Alt + F11",
            "2. Ou: Menu Desenvolvedor → Visual Basic",
            "3. Nota: Ativar aba Desenvolvedor se não visível",
        ]),
        ("Macro Simples", [
            "Sub saudacao()",
            "    MsgBox 'Olá, mundo!'",
            "End Sub",
            "",
            "Para executar: Alt + F8 → Selecionar macro",
        ]),
    ], header_fill, header_font, section_fill, section_font, border)

    # Salvar workbook
    output_file = "/home/user/SME/Tutorial_Dashboard.xlsx"
    wb.save(output_file)
    print(f"✓ Planilha criada com sucesso: {output_file}")
    print(f"✓ Total de abas: {len(wb.sheetnames)}")
    print(f"✓ Abas criadas: {', '.join(wb.sheetnames)}")


def _create_tutorial_tab(ws, title, sections, header_fill, header_font, section_fill, section_font, border):
    """Função auxiliar para criar abas de tutorial com formatação consistente"""

    # Título da aba
    ws['A1'] = title.upper()
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:B1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    current_row = 3

    # Adicionar seções
    for section_title, content in sections:
        # Título da seção
        ws.cell(row=current_row, column=1).value = section_title
        ws.cell(row=current_row, column=1).font = section_font
        ws.cell(row=current_row, column=1).fill = section_fill
        ws.cell(row=current_row, column=1).border = border
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws.row_dimensions[current_row].height = 18

        current_row += 1

        # Conteúdo
        if isinstance(content, str):
            # Conteúdo é texto simples
            ws.cell(row=current_row, column=1).value = content
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.cell(row=current_row, column=1).border = border
            ws.merge_cells(f'A{current_row}:B{current_row}')
            ws.row_dimensions[current_row].height = 30
            current_row += 1
        else:
            # Conteúdo é lista de pontos
            for point in content:
                ws.cell(row=current_row, column=1).value = point
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                ws.cell(row=current_row, column=1).border = border
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws.row_dimensions[current_row].height = 20
                current_row += 1

        # Espaço entre seções
        current_row += 1

    # Ajustar larguras de coluna
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 50


if __name__ == "__main__":
    create_tutorial_dashboard()
